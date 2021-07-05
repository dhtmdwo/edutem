from openpyxl import load_workbook
from grammarbot import grammarBot
from language_tool import correction
import time
from ginger import Ginger_tool
xl_file = 'sample_sentences(grammarly).xlsx'

success = 0  # 성공
total = 0  # 총
wb = load_workbook(filename=xl_file)
#print('sheet count: ', len(wb.sheetnames))
ws = wb['테스트 문장']

get_cells = ws['B3':'B83']  # 문제 셀 범위 입력하기
total = len(get_cells)
corrected = []

start = time.time()
for row in get_cells:
    for cell in row:
        # print(cell.value)
        if cell.value is None:
            continue
        # print(cell.value)
        # corrected.append(grammarBot(cell.value))
        # sent = correction(cell.value)->language tool
        sent = Ginger_tool(cell.value)
        # print(sent)
        corrected.append(sent)


get_cells = ws['D3':'D83']  # 정답 셀 범위 입력하기
count = 0
for row in get_cells:
    for cell in row:
        if cell.value is None:
            continue
        else:
            if corrected[count] == cell.value:
                print('Corrected: '+corrected[count]+' Answer: '+cell.value)
                success += 1
            print('Wrong: '+corrected[count]+' Answer: '+cell.value)
        count += 1
possibility = success/total
print('정답률: '+str(possibility))
print('소요 시간' + str(round((time.time()-start), 2)) + 's')
